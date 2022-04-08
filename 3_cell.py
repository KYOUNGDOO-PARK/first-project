from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

#A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

wb.save("sample.xlsx")

